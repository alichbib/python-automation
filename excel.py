from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# #work book(excel file)
# #xlsx the extention
# wb = load_workbook('test/Book1.xlsx')

# #active work sheet
# ws = wb.active

# #printing the value A1
# print(ws['A1'].value)

# #excel file should be closed
# #updating value A2
# ws['A2'].value = "Test" 
# wb.save('test/Book1.xlsx')

# ws = wb['sheet1']

# wb.create_sheet('Grades22')


# ws.append(['start'])
# ws.append(['ali', 'is', 'a', 'programmer'])
# ws.append(["end"])
# wb.save('test/Grades.xlsx')


wb = load_workbook('test/Grades.xlsx')
ws = wb.active
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)
        


wb.save('test/Grades.xlsx')